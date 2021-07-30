import openpyxl

wb = openpyxl.load_workbook('sample.xlsx')

ws = wb.active

# iter_cols()メソッドで読み込み
for col in ws.iter_cols(min_row=1, min_col=1):
    for cell in col:
        print(cell.value)

# iter_rows()メソッドで読み込み
for row in ws.iter_rows(min_row=1, min_col=1):
    for cell in row:
        print(cell.value)

### 定刻よりも早く出発（到着）していれば◯、していなければ✕を表す列を追加する ###

# 各列(D, F)の出発時間を取得する
STD = []
for col in ws.iter_cols(min_row=2, min_col=4, max_col=4):
    for cell in col:
        STD.append(cell.value)

ATD = []
for col in ws.iter_cols(min_row=2, min_col=6, max_col=6):
    for cell in col:
        ATD.append(cell.value)

for i, k in zip(STD, ATD):
    print(i, k)

# 時間を判定する
DEP = []
for i, k in zip(STD, ATD):
    if i >= k:
        DEP.append('◯')
    else:
        DEP.append('✕')

# 判定結果を新しい列に書き込む
ws['I1'] = '定時出発'
for col in ws.iter_cols(min_row=2, min_col=9, 
                        max_row=len(DEP)+1, max_col=9):
    for cell, i in zip(col, DEP):
        cell.value = i

wb.save('sample.xlsx')

# 各列(E, G)の到着時間を取得する
STA = []
for col in ws.iter_cols(min_row=2, min_col=5, max_col=5):
    for cell in col:
        STA.append(cell.value)

ATA = []
for col in ws.iter_cols(min_row=2, min_col=7, max_col=7):
    for cell in col:
        ATA.append(cell.value)

# 時間を判定する
ARR = []
for i, k in zip(STA, ATA):
    if i >= k:
        ARR.append('◯')
    else:
        ARR.append('✕')

# 判定結果を新しい列に書き込む
ws['J1'] = '定時到着'
for col in ws.iter_cols(min_row=2, min_col=10, 
                        max_row=len(ARR)+1, max_col=10):
    for cell, i in zip(col, ARR):
        cell.value = i

wb.save('sample.xlsx')

### 乗車率を表す列を作成する ###

# 最大席数を45と仮定して、各便の乗車率を計算する（乗客数÷45）
ws['K1'] = '乗車率'

LOAD_FACTOR = []
for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
    for cell in col:
        lf = cell.value / 45
        LOAD_FACTOR.append(lf)

# openpyxlの機能を使って%表記にする
from openpyxl.styles import numbers

for col in ws.iter_cols(min_row=2, min_col=11,
                        max_row=len(LOAD_FACTOR)+1, max_col=11):
    for cell, i in zip(col, LOAD_FACTOR):
        cell.value = i

for row in ws.iter_rows(min_row=2, min_col=11,
                        max_row=len(LOAD_FACTOR)+1, max_col=11):
    for cell in row:
        cell.number_format = numbers.FORMAT_PERCENTAGE
        
wb.save('sample.xlsx')

# summaryシートを追加する
ws2 = wb.create_sheet('summary')
print(wb.sheetnames)

# 定時出発率、定時到着率、平均乗車率をchartシートに反映する
ws2['A1'] = '定時出発率'
dep_result = []
for col in ws.iter_cols(min_row=2, min_col=9, max_col=9):
    for cell in col:
        dep_result.append(cell.value)

dep_rate = dep_result.count('◯') / len(dep_result)
ws2['A2'] = '{:.0%}'.format(dep_rate)

ws2['B1'] = '定時到着率'
arr_result = []
for col in ws.iter_cols(min_row=2, min_col=10, max_col=10):
    for cell in col:
        arr_result.append(cell.value)

arr_rate = arr_result.count('◯') / len(arr_result)
ws2['B2'] = '{:.0%}'.format(arr_rate)

ws2['C1'] = '平均乗車率'
lf_result = 0
for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
    for cell in col:
        lf_result += cell.value / 45
lf_result = lf_result / len(arr_result)
ws2['C2'] = '{:.0%}'.format(lf_result)

wb.save('sample.xlsx')

### 運行数の棒グラフを作成する ###

# 区間ごとの運行数を作成する
route_num = {}
for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
    for cell in col:
        if cell.value in route_num:
            route_num[cell.value] += 1
        else:
            route_num[cell.value] = 1

for i,k in route_num.items():
    print(i, k)

# summaryシートにデータを反映する
route = [i for i in route_num.keys()]
route.insert(0, '区間')

num = [i for i in route_num.values()]
num.insert(0, '運行数')

for col in ws2.iter_cols(min_row=1, min_col=5,
                         max_row=len(route), max_col=5):
    for cell, r in zip(col, route):
        cell.value = r

for col in ws2.iter_cols(min_row=1, min_col=6, 
                         max_row=len(num), max_col=6):
    for cell, n in zip(col, num):
        cell.value = n

wb.save('sample.xlsx')

### グラフを挿入する ###

# 必要な機能をimportする
from openpyxl.chart import BarChart, Reference, Series
    
# 棒グラフを初期化する
chart = BarChart()
chart.title = '区間別運行数'

# 描画するための値を作成
plot_value = Reference(ws2, min_row=1, min_col=6, max_row=5, max_col=6)
chart.add_data(plot_value, titles_from_data=True)
# X軸のラベルを作成
x = Reference(ws2, min_row=2, min_col=5, max_row=5, max_col=5)
chart.set_categories(x)
    
ws2.add_chart(chart, 'E6')

wb.save('sample.xlsx')

# LFの折れ線グラフを作成する
from openpyxl.chart import LineChart

# 折れ線グラフを初期化する
linechart = LineChart()
linechart.title = 'LF推移'

# 描画するための値を作成
plot_value = Reference(ws, min_row=1, min_col=11, max_row=61, max_col=11)
linechart.add_data(plot_value, titles_from_data=True)

# X軸のラベルを作成
x = Reference(ws, min_row=2, min_col=1, max_row=61, max_col=1)
linechart.set_categories(x)

ws2.add_chart(linechart, 'N6')

wb.save('sample.xlsx')

# 区間ごとのシートを作成する
# グラフ作成時に作成したroute変数を使用する
del route[0]

for i in route:
    wb.create_sheet(i)

print(wb.sheetnames)

# for分で区間のシートを取得して列名を書き込む
for i in range(2, 6):
    worksheet = wb.worksheets[i]

    worksheet['A1'] = '日付'
    worksheet['B1'] = '便名'
    worksheet['C1'] = '区間'
    worksheet['D1'] = '予定出発時間'
    worksheet['E1'] = '予定到着時間'
    worksheet['F1'] = '出発時間'
    worksheet['G1'] = '到着時間'
    worksheet['H1'] = '乗客数'
    worksheet['I1'] = '定時出発'
    worksheet['J1'] = '定時到着'
    worksheet['K1'] = '乗車率'

wb.save('sample.xlsx')

for i in range(2, 6):
    worksheet = wb.worksheets[i]

    # 日付列の書き込み
    DATE = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                DATE.append(ws['A' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=1, 
                                   max_row=len(DATE)+1, max_col=1):
        for cell, j in zip(col, DATE):
            cell.value = j

    for col in worksheet.iter_cols(min_row=2, min_col=1,
                                   max_row=len(DATE)+1, max_col=1):
        for cell in col:
            cell.number_format = 'YYYY/MM/DD'

wb.save('sample.xlsx')

for i in range(2, 6):
    worksheet = wb.worksheets[i]

    # 便名の書き込み
    FLIGHT_NAME = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                FLIGHT_NAME.append(ws['B' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=2, 
                                   max_row=len(FLIGHT_NAME)+1, max_col=2):
        for cell, j in zip(col, FLIGHT_NAME):
            cell.value = j

    # 区間の書き込み
    ROUTE = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                ROUTE.append(ws['C' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=3, 
                                   max_row=len(ROUTE)+1, max_col=3):
        for cell, j in zip(col, ROUTE):
            cell.value = j

wb.save('sample.xlsx')

for i in range(2, 6):
    worksheet = wb.worksheets[i]

    # 予定出発時間の書き込み
    SKD_DEP_TIME = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                SKD_DEP_TIME.append(ws['D' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=4, 
                                   max_row=len(SKD_DEP_TIME)+1, max_col=4):
        for cell, j in zip(col, SKD_DEP_TIME):
            cell.value = j

    # 予定到着時間の書き込み
    SKD_ARR_TIME = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                SKD_ARR_TIME.append(ws['E' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=5,
                                   max_row=len(SKD_ARR_TIME)+1, max_col=5):
        for cell, j in zip(col, SKD_ARR_TIME):
            cell.value = j

    # 出発時間の書き込み
    ACT_DEP_TIME = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                ACT_DEP_TIME.append(ws['F' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=6,
                                   max_row=len(ACT_DEP_TIME)+1, max_col=6):
        for cell, j in zip(col, ACT_DEP_TIME):
            cell.value = j

    # 到着時間の書き込み
    ACT_ARR_TIME = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                ACT_ARR_TIME.append(ws['G' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=7,
                                   max_row=len(ACT_ARR_TIME)+1, max_col=7):
        for cell, j in zip(col, ACT_ARR_TIME):
            cell.value = j

    # 時間表記をHH:MMにする
    for col in worksheet.iter_cols(min_row=2, min_col=4, max_col=7):
        for cell in col:
            cell.number_format = 'HH:MM'

wb.save('sample.xlsx')

for i in range(2, 6):
    worksheet = wb.worksheets[i]

    # 乗客数の書き込み
    PAX = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                PAX.append(ws['H' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=8,
                                   max_row=len(PAX)+1, max_col=8):
        for cell, j in zip(col, PAX):
            cell.value = j

    # 定時出発（◯✕）の書き込み
    DEP_marubatu = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                DEP_marubatu.append(ws['I' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=9,
                                   max_row=len(DEP_marubatu)+1, max_col=9):
        for cell, j in zip(col, DEP_marubatu):
            cell.value = j

    # 定時到着（◯✕）の書き込み
    ARR_marubatu = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                ARR_marubatu.append(ws['J' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=10,
                                   max_row=len(ARR_marubatu)+1, max_col=10):
        for cell, j in zip(col, ARR_marubatu):
            cell.value = j

    # 乗車率の書き込み
    LF = []
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell, num in zip(col, range(2, len(DEP)+2)):
            if cell.value == wb.sheetnames[i]:
                LF.append(ws['K' + str(num)].value)

    for col in worksheet.iter_cols(min_row=2, min_col=11,
                                   max_row=len(LF)+1, max_col=11):
        for cell, j in zip(col, LF):
            cell.value = j

    for col in worksheet.iter_cols(min_row=2, min_col=11,
                                   max_row=len(LF)+1, max_col=11):
        for cell in col:
            cell.number_format = numbers.FORMAT_PERCENTAGE

wb.save('sample.xlsx')

# グラフを挿入する
for i in range(2, 6):
    worksheet = wb.worksheets[i]

    # 折れ線グラフを初期化する
    linecharts = LineChart()
    linecharts.title = 'LF推移'

    # 最終行を取得
    max_row = worksheet.max_row

    # 描画するための値を作成
    plot_value = Reference(worksheet, min_row=1, min_col=11,
                                      max_row=max_row, max_col=11)
    linecharts.add_data(plot_value, titles_from_data=True)

    # X軸のラベルを作成
    x = Reference(worksheet, min_row=2, min_col=1,
                             max_row=max_row, max_col=1)
    linecharts.set_categories(x)

    worksheet.add_chart(linecharts, 'M2')

wb.save('sample.xlsx')

# おまけ（フォントを変更する）
from openpyxl.styles import Font

font = Font(name='meiryo', size=10)

for i in range(0, len(wb.sheetnames)):
    worksheet = wb.worksheets[i]
    
    for col in worksheet.iter_cols(min_row=1, min_col=1):
        for cell in col:
            worksheet[cell.coordinate].font = font

wb.save('sample.xlsx')
