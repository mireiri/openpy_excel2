import openpyxl

wb = openpyxl.load_workbook('sample.xlsx')

ws = wb.active

# iter_cols()メソッドで読み込み
for col in ws.iter_cols(min_row=1, min_col=1):
    for cell in col:
        print(cell.value)

# iter_rows()メソッドで読み込み
for row in ws.iter_rows(min_row=1, min_col=1):
    for cell in row:
        print(cell.value)

### 定刻よりも早く出発（到着）していれば◯、していなければ✕を表す列を追加する ###

# 各列(D, F)の出発時間を取得する
STD = []
for col in ws.iter_cols(min_row=2, min_col=4, max_col=4):
    for cell in col:
        STD.append(cell.value)

ATD = []
for col in ws.iter_cols(min_row=2, min_col=6, max_col=6):
    for cell in col:
        ATD.append(cell.value)

for i, k in zip(STD, ATD):
    print(i, k)

# 時間を判定する
DEP = []
for i, k in zip(STD, ATD):
    if i >= k:
        DEP.append('◯')
    else:
        DEP.append('✕')

# 判定結果を新しい列に書き込む
ws['I1'] = '定時出発'
for col in ws.iter_cols(min_row=2, min_col=9, 
                        max_row=len(DEP)+1, max_col=9):
    for cell, i in zip(col, DEP):
        cell.value = i

wb.save('sample.xlsx')

# 各列(E, G)の到着時間を取得する
STA = []
for col in ws.iter_cols(min_row=2, min_col=5, max_col=5):
    for cell in col:
        STA.append(cell.value)

ATA = []
for col in ws.iter_cols(min_row=2, min_col=7, max_col=7):
    for cell in col:
        ATA.append(cell.value)

# 時間を判定する
ARR = []
for i, k in zip(STA, ATA):
    if i >= k:
        ARR.append('◯')
    else:
        ARR.append('✕')

# 判定結果を新しい列に書き込む
ws['J1'] = '定時到着'
for col in ws.iter_cols(min_row=2, min_col=10, 
                        max_row=len(ARR)+1, max_col=10):
    for cell, i in zip(col, ARR):
        cell.value = i

wb.save('sample.xlsx')

### 乗車率を表す列を作成する ###

# 最大席数を45と仮定して、各便の乗車率を計算する（乗客数÷45）
ws['K1'] = '乗車率'

LOAD_FACTOR = []
for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
    for cell in col:
        lf = cell.value / 45
        LOAD_FACTOR.append(lf)

# openpyxlの機能を使って%表記にする
from openpyxl.styles import numbers

for col in ws.iter_cols(min_row=2, min_col=11,
                        max_row=len(LOAD_FACTOR)+1, max_col=11):
    for cell, i in zip(col, LOAD_FACTOR):
        cell.value = i

for row in ws.iter_rows(min_row=2, min_col=11,
                        max_row=len(LOAD_FACTOR)+1, max_col=11):
    for cell in row:
        cell.number_format = numbers.FORMAT_PERCENTAGE
        
wb.save('sample.xlsx')

# summaryシートを追加する
ws2 = wb.create_sheet('summary')

print(wb.sheetnames)

# 定時出発率、定時到着率、平均乗車率をchartシートに反映する
ws2['A1'] = '定時出発率'
dep_result = []
for col in ws.iter_cols(min_row=2, min_col=9, max_col=9):
    for cell in col:
        dep_result.append(cell.value)

dep_rate = dep_result.count('◯') / len(dep_result)
ws2['A2'] = '{:.0%}'.format(dep_rate)

ws2['B1'] = '定時到着率'
arr_result = []
for col in ws.iter_cols(min_row=2, min_col=10, max_col=10):
    for cell in col:
        arr_result.append(cell.value)

arr_rate = arr_result.count('◯') / len(arr_result)
ws2['B2'] = '{:.0%}'.format(arr_rate)

ws2['C1'] = '平均乗車率'
lf_result = 0
for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
    for cell in col:
        lf_result += cell.value / 45
lf_result = lf_result / len(arr_result)
ws2['C2'] = '{:.0%}'.format(lf_result)

wb.save('sample.xlsx')

### 運行数の棒グラフを作成する ###

# 区間ごとの運行数を作成する
route_num = {}
for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
    for cell in col:
        if cell.value in route_num:
            route_num[cell.value] += 1
        else:
            route_num[cell.value] = 1

for i,k in route_num.items():
    print(i, k)

# summaryシートにデータを反映する
route = [i for i in route_num.keys()]
route.insert(0, '区間')

num = [i for i in route_num.values()]
num.insert(0, '運行数')

for col in ws2.iter_cols(min_row=1, min_col=5,
                         max_row=len(route), max_col=5):
    for cell, r in zip(col, route):
        cell.value = r

for col in ws2.iter_cols(min_row=1, min_col=6, 
                         max_row=len(num), max_col=6):
    for cell, n in zip(col, num):
        cell.value = n

wb.save('sample.xlsx')

### グラフを挿入する ###

# 必要な機能をimportする
from openpyxl.chart import BarChart, Reference, Series
    
# 棒グラフを初期化する
chart = BarChart()
chart.title = '区間別運行数'

# 描画するための値を作成
plot_value = Reference(ws2, min_row=1, min_col=6, max_row=5, max_col=6)
chart.add_data(plot_value, titles_from_data=True)
# X軸のラベルを作成
x = Reference(ws2, min_row=2, min_col=5, max_row=5, max_col=5)
chart.set_categories(x)
    
ws2.add_chart(chart, 'E6')

wb.save('sample.xlsx')

# LFの折れ線グラフを作成する
from openpyxl.chart import LineChart

# 折れ線グラフを初期化する
linechart = LineChart()
linechart.title = 'LF推移'

# 描画するための値を作成
plot_value = Reference(ws, min_row=1, min_col=11, max_row=61, max_col=11)
linechart.add_data(plot_value, titles_from_data=True)

# X軸のラベルを作成
x = Reference(ws, min_row=2, min_col=1, max_row=61, max_col=1)
linechart.set_categories(x)

ws2.add_chart(linechart, 'N6')

wb.save('sample.xlsx')